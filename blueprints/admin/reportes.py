"""
Exportación de datos del panel admin a Excel.

- ``admin_exportar`` → genera un archivo .xlsx con estilos de la marca para
                       usuarios, carreras, preguntas o noticias (requisito TFI:
                       exportación de datos), respetando los filtros activos.
"""

import io
from datetime import datetime

from flask import Blueprint, Response, redirect, request, url_for

from core.decoradores import requiere_admin
from database_handler import obtener_db

bp = Blueprint('admin_reportes', __name__)


@bp.route('/admin/exportar/<entidad>')
@requiere_admin
def admin_exportar(entidad):
    """Exporta datos del panel a Excel (requisito TFI: exportación de datos)."""
    db = obtener_db()
    cursor = db.cursor(dictionary=True)

    configs = {
        'usuarios': {
            'tabla': 'usuarios',
            'nombre_archivo': 'usuarios',
            'columnas': ['id', 'nombre', 'email', 'rol', 'created_at'],
            'titulos': ['ID', 'Nombre', 'Email', 'Rol', 'Fecha de registro'],
        },
        'carreras': {
            'tabla': 'carreras',
            'nombre_archivo': 'carreras',
            'columnas': ['id', 'nombre', 'area_profesional', 'descripcion', 'a_que_se_dedica'],
            'titulos': ['ID', 'Nombre', 'Área profesional', 'Descripción', '¿A qué se dedica?'],
        },
        'preguntas': {
            'tabla': 'preguntas',
            'nombre_archivo': 'preguntas',
            'columnas': ['id', 'texto_pregunta', 'area_profesional'],
            'titulos': ['ID', 'Pregunta', 'Área profesional'],
        },
        'noticias': {
            'tabla': 'noticias',
            'nombre_archivo': 'noticias',
            'columnas': ['id', 'titulo', 'categoria', 'fuente', 'fecha', 'link'],
            'titulos': ['ID', 'Título', 'Categoría', 'Fuente', 'Fecha', 'Enlace'],
        },
    }

    config = configs.get(entidad)
    if not config:
        return redirect(url_for('admin.admin_dashboard'))

    columnas = config['columnas']

    # Aplica los filtros activos del panel a la exportación.
    where = []
    params = []
    if entidad == 'usuarios':
        f_nombre = request.args.get('nombre', '').strip()
        f_email = request.args.get('email', '').strip()
        f_fecha = request.args.get('fecha', '').strip()
        if f_nombre:
            where.append("(nombre LIKE %s OR apellido LIKE %s)")
            params.extend(['%' + f_nombre + '%'] * 2)
        if f_email:
            where.append("email LIKE %s")
            params.append('%' + f_email + '%')
        if f_fecha:
            where.append("DATE(created_at) = %s")
            params.append(f_fecha)
    elif entidad == 'noticias':
        f_fuente = request.args.get('fuente', 'todas')
        f_categoria = request.args.get('categoria', 'todas')
        busqueda = request.args.get('q', '').strip()
        if f_fuente != 'todas':
            where.append("fuente = %s")
            params.append(f_fuente)
        if f_categoria != 'todas':
            where.append("categoria = %s")
            params.append(f_categoria)
        if busqueda:
            where.append("(titulo LIKE %s OR descripcion LIKE %s OR fuente LIKE %s)")
            params.extend([f"%{busqueda}%", f"%{busqueda}%", f"%{busqueda}%"])

    sql = f"SELECT {', '.join(columnas)} FROM {config['tabla']}"
    if where:
        sql += " WHERE " + " AND ".join(where)
    cursor.execute(sql, params)
    filas = cursor.fetchall()

    # --- Generación del archivo Excel profesional ---
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = config['titulos'][1] if len(config['titulos']) > 1 else 'Datos'

    # Colores de la marca.
    AZUL_PRIMARIO = '2F8EAB'
    AZUL_OSCURO = '142B38'
    GRIS_BORDE = 'D9E2E8'
    GRIS_FILA = 'F4F8FA'

    thin_border = Border(
        left=Side(style='thin', color=GRIS_BORDE),
        right=Side(style='thin', color=GRIS_BORDE),
        top=Side(style='thin', color=GRIS_BORDE),
        bottom=Side(style='thin', color=GRIS_BORDE),
    )

    # Encabezados: fondo azul oscuro + texto blanco en negrita.
    for j, titulo in enumerate(config['titulos'], start=1):
        cell = ws.cell(row=1, column=j, value=titulo)
        cell.font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
        cell.fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[1].height = 26

    # Datos.
    for i, fila in enumerate(filas, start=2):
        for j, col in enumerate(columnas, start=1):
            valor = fila.get(col, '')
            if hasattr(valor, 'strftime'):
                valor = valor.strftime('%d/%m/%Y %H:%M')
            cell = ws.cell(row=i, column=j, value=valor)
            cell.font = Font(size=11, name='Calibri')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
        # Bandas alternadas suaves para mejor lectura.
        if i % 2 == 0:
            for j in range(1, len(columnas) + 1):
                ws.cell(row=i, column=j).fill = PatternFill(
                    start_color=GRIS_FILA, end_color=GRIS_FILA, fill_type='solid')

    # Ajustar ancho de columnas según el contenido.
    for j, col in enumerate(columnas, start=1):
        max_len = len(str(config['titulos'][j - 1]))
        for fila in filas:
            val = fila.get(col, '')
            if hasattr(val, 'strftime'):
                val = val.strftime('%d/%m/%Y %H:%M')
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 4, 60)

    # Autofiltro + panel congelado en el encabezado.
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{len(filas) + 1}"
    ws.freeze_panes = 'A2'

    # Hoja de resumen con metadatos.
    ws_meta = wb.create_sheet('Resumen')
    ws_meta.sheet_view.showGridLines = False
    ws_meta.column_dimensions['A'].width = 32
    ws_meta.column_dimensions['B'].width = 60

    ws_meta['A1'] = 'Futuro 360 — Panel de Administración'
    ws_meta['A1'].font = Font(bold=True, size=14, color=AZUL_OSCURO)
    ws_meta['A2'] = 'Exportación de datos'
    ws_meta['A2'].font = Font(size=12, color=AZUL_PRIMARIO)
    ws_meta['A4'] = 'Entidad'
    ws_meta['B4'] = config['titulos'][0]
    ws_meta['A5'] = 'Total de registros'
    ws_meta['B5'] = len(filas)
    ws_meta['A6'] = 'Fecha de exportación'
    ws_meta['B6'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    for fila in [4, 5, 6]:
        ws_meta.cell(row=fila, column=1).font = Font(bold=True, color=AZUL_OSCURO)
        ws_meta.cell(row=fila, column=2).font = Font(color='445A6B')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha = datetime.now().strftime('%Y-%m-%d')
    nombre = f"{config['nombre_archivo']}_{fecha}.xlsx"

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={nombre}'}
    )
